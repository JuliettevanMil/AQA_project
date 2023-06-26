import numpy as np
from functions_dwave import find_solution
from check_and_correct import check_instances
from openpyxl import load_workbook

def add_list_to_datafile(data_list1, data_list2, file_path, sheet_name):
    # Load the existing workbook
    workbook = load_workbook(file_path)

    # Select the specified sheet
    sheet = workbook[sheet_name]

    # Find the next available row
    next_row = sheet.max_row + 1

    sheet.cell(row=next_row, column=1, value=alpha)
    sheet.cell(row=next_row, column=2, value=beta)

    # Iterate over the list and write the values to the sheet
    for column, data1 in enumerate(data_list1, start=3):
        sheet.cell(row=next_row, column=column, value=data1)
    for column, data2 in enumerate(data_list2, start=3):    
        sheet.cell(row=next_row+1,column=column, value=data2)

    # Save the modified workbook
    workbook.save(file_path)

def create_alternating_matrix(d):
    input_matrix = np.zeros((d, d), dtype=int)
    for i in range(d):
        for j in range(d):
            if (i + j) % 2 == 0:
                input_matrix[i][j] = 1
    return input_matrix

# d = 4
alpha = 1  # reward
# beta = 1  # penalty
beta_lst = [alpha/2, alpha, 3*alpha/2, 2*alpha, 5*alpha/2, 3*alpha, 4*alpha, 5*alpha, 6*alpha]
# beta_lst = [7*alpha, 8*alpha, 9*alpha, 10*alpha]
runs = 10

data_file = 'Results_threshold.xlsx'

for d in range(3,10):
    print(d)
    input_matrix = create_alternating_matrix(d)
    for beta in beta_lst:
        print(beta)
        num_correct_lst = np.zeros(runs)
        num_false_lst = np.zeros(runs)

        for i in range(runs):
            solution_dict, lowest_energy = find_solution(
                input_matrix,
                alpha,
                beta,
                cluster_dim="col",  # choose from ["col", "row"]
                scale="treshold",  # choose from ["linear","treshold","id"]
                show_results=False,  # choose from [True, False]
            )
            # print(solution_dict)

            checked, correction = check_instances(solution_dict, fix=True)
            if beta >= 1:
                print(checked)
            # print(corrections)
            # checked_corrections = check_instances(corrections, fix=False)
            # print(checked_corrections)

            num_correct = 0
            for sample in checked:
                if checked[sample][-1] == True:
                    num_correct += 1
            
            num_correct_lst[i] = num_correct
            num_false_lst[i] = len(checked) - num_correct

        print(np.mean(num_correct_lst), np.mean(num_false_lst))

        print("correct outputs: ", num_correct_lst)
        print("false outputs: ", num_false_lst)

        add_list_to_datafile(num_correct_lst, num_false_lst, data_file, '{}'.format(d))